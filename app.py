import streamlit as st
import pandas as pd
import plotly.express as px
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import PieChart, BarChart, Reference
import numpy as np
from pathlib import Path

# -----------------------
# CONFIG
# -----------------------
st.set_page_config(
    page_title="Dashboard PPNS Otomatis",
    layout="wide",
    page_icon="📊"
)

# -----------------------
# LOAD EXTERNAL CSS
# -----------------------
css_path = Path("style.css")
if css_path.exists():
    with open(css_path) as f:
        st.markdown(f"<style>{f.read()}</style>", unsafe_allow_html=True)

st.markdown(
    """
    <div style="
        background: linear-gradient(180deg, #0077B6 0%, #005F99 100%);
        border-radius: 10px;
        padding: 15px 0 25px 0;
        margin-bottom: 25px;
        box-shadow: 0 4px 20px rgba(0,0,0,0.1);
    ">
    """,
    unsafe_allow_html=True
)

# -----------------------
# HEADER NAVY
# -----------------------
col1, col2, col3 = st.columns([1, 5, 1])

with col1:
    if Path("logo.png").exists():
        st.image("logo.png", width=90)

with col2:
    st.markdown("""
    <div style='text-align:center;'>
        <div style='font-size:28px; font-weight:700; color:#00BFFF; margin-bottom:4px;'>
            📘 Dashboard Direktorat Penanganan Pelanggaran
        </div>
        <div style='font-size:18px; font-weight:600; color:#00BFFF; margin-bottom:4px;'>
            Kementerian Kelautan dan Perikanan Republik Indonesia — Data PPNS
        </div>
    </div>
    """, unsafe_allow_html=True)

with col3:
    if Path("logo_kkp.png").exists():
        st.image("logo_kkp.png", width=90)

st.markdown(
    """
    <hr style='margin-top:10px; margin-bottom:25px; border: 1px solid #00BFFF; opacity:0.5;'/>
    """,
    unsafe_allow_html=True
)


# -----------------------
# DARK / LIGHT MODE TOGGLE (NAVY + GOLD PREMIUM)
# -----------------------
st.markdown("<br>", unsafe_allow_html=True)
dark_mode = st.toggle("🌙 Mode Gelap", key="darkmode_toggle")

if dark_mode:
    # 🌙 Dark Mode: navy tua + gold mewah
    st.markdown("""
    <style>
    :root {
        --bg-color: #0A0F2E;
        --card-bg: #1B2545;
        --text-color: #FDF6E3;
        --accent: #D4AF37; /* gold premium */
        --accent-hover: #FFD700;
        --navy: #1B2545;
        --sidebar-bg: linear-gradient(to bottom, #0A0F2E, #1B2545);
        --card-shadow: rgba(212, 175, 55, 0.3);
        --transition-speed: 0.3s;
    }
    /* Body & main container */
    body, [class*="stAppViewContainer"], [data-testid="stHeader"] {
        background-color: var(--bg-color) !important;
        color: var(--text-color) !important;
        transition: background-color var(--transition-speed), color var(--transition-speed);
    }
    /* Sidebar */
    .stSidebar {
        background: var(--sidebar-bg) !important;
        color: var(--text-color) !important;
        transition: background var(--transition-speed), color var(--transition-speed);
    }
    /* Text & headings */
    div, p, span, label, h1, h2, h3, h4, h5 {
        color: var(--text-color) !important;
        transition: color var(--transition-speed);
        font-family: "Segoe UI", sans-serif;
    }
    h1, .stMarkdown h1 {
        color: var(--accent) !important;
        font-weight: 700;
        letter-spacing: 1px;
    }
    /* Buttons */
    .stButton button {
        background-color: var(--accent) !important;
        color: var(--bg-color) !important;
        border-radius: 12px;
        border: none;
        padding: 0.55em 1.5em;
        font-weight: 600;
        transition: background-color var(--transition-speed), color var(--transition-speed), transform var(--transition-speed);
        box-shadow: 0 4px 12px var(--card-shadow);
    }
    .stButton button:hover {
        background-color: var(--accent-hover) !important;
        color: #0A0F2E !important;
        transform: translateY(-2px);
    }
    /* Cards */
    [class*="stCard"] {
        background-color: var(--card-bg) !important;
        border-radius: 12px;
        padding: 1rem;
        box-shadow: 0 6px 18px var(--card-shadow);
        transition: background-color var(--transition-speed), box-shadow var(--transition-speed);
    }
    /* Table */
    thead tr th {
        background-color: var(--navy) !important;
        color: var(--accent) !important;
        font-weight: 600;
    }
    tbody tr td {
        color: var(--text-color) !important;
    }
    </style>
    """, unsafe_allow_html=True)
else:
    # 🌞 Light Mode: putih + navy + gold lembut
    st.markdown("""
    <style>
    :root {
        --bg-color: #FFFFFF;
        --card-bg: #F2F5F9;
        --text-color: #1B263B;
        --accent: #C6A664; /* gold lembut */
        --accent-hover: #D4AF37;
        --navy: #0D47A1;
        --sidebar-bg: linear-gradient(to bottom, #003366, #0055A5);
        --card-shadow: rgba(198, 166, 100, 0.2);
        --transition-speed: 0.3s;
    }
    body, [class*="stAppViewContainer"], [data-testid="stHeader"] {
        background-color: var(--bg-color) !important;
        color: var(--text-color) !important;
        transition: background-color var(--transition-speed), color var(--transition-speed);
    }
    .stSidebar {
        background: var(--sidebar-bg) !important;
        color: #FFFFFF !important;
        transition: background var(--transition-speed), color var(--transition-speed);
    }
    div, p, span, label, h1, h2, h3, h4, h5 {
        color: var(--text-color) !important;
        transition: color var(--transition-speed);
        font-family: "Segoe UI", sans-serif;
    }
    h1, .stMarkdown h1 {
        color: var(--navy) !important;
        font-weight: 700;
        letter-spacing: 0.5px;
    }
    .stButton button {
        background-color: var(--navy) !important;
        color: #FFFFFF !important;
        border-radius: 12px;
        border: 1px solid var(--accent);
        padding: 0.55em 1.5em;
        font-weight: 600;
        transition: background-color var(--transition-speed), color var(--transition-speed), transform var(--transition-speed);
        box-shadow: 0 4px 12px var(--card-shadow);
    }
    .stButton button:hover {
        background-color: var(--accent-hover) !important;
        color: #FFFFFF !important;
        transform: translateY(-2px);
    }
    [class*="stCard"] {
        background-color: var(--card-bg) !important;
        border-radius: 12px;
        padding: 1rem;
        box-shadow: 0 6px 18px var(--card-shadow);
        transition: background-color var(--transition-speed), box-shadow var(--transition-speed);
    }
    thead tr th {
        background-color: var(--navy) !important;
        color: var(--accent) !important;
        font-weight: 600;
    }
    tbody tr td {
        color: var(--text-color) !important;
    }
    </style>
    """, unsafe_allow_html=True)


# -----------------------
# HELPERS
# -----------------------
UNWANTED_LABELS = {
    "-", "", "nan", "none", "unknown", "n/a", "na", "unk", "other", "tidak diketahui", "null", "unknowns", "not available"
}

def is_unwanted_label(s):
    if s is None:
        return True
    s2 = str(s).strip().lower()
    return s2 in UNWANTED_LABELS

def clean_series_for_vc(series: pd.Series) -> pd.Series:
    try:
        s = series.dropna().astype(str).map(lambda x: x.strip())
    except Exception:
        s = series.dropna().astype(str).map(str)
    s = s[s != ""]
    s = s[~s.str.lower().isin(UNWANTED_LABELS)]
    return s

def detect_header_row(raw_df, keywords=None):
    if keywords is None:
        keywords = ["nama", "nip", "gender", "jenis kelamin", "jabatan", "pangkat", "instansi", "divisi"]
    for i, row in raw_df.iterrows():
        row_str = [str(x).strip().lower() for x in row.tolist()]
        for kw in keywords:
            if any(kw in cell for cell in row_str if cell and cell != 'nan'):
                return i
    return 0

def _safe_value_for_excel(x):
    if x is None or x is pd.NA:
        return None
    try:
        if isinstance(x, float) and np.isnan(x):
            return None
    except Exception:
        pass
    if isinstance(x, (np.integer,)):
        return int(x)
    if isinstance(x, (np.floating,)):
        return float(x)
    if isinstance(x, (np.bool_, bool)):
        return bool(x)
    if isinstance(x, str):
        return x
    return x

def safe_for_excel_df(df: pd.DataFrame) -> pd.DataFrame:
    df2 = df.copy().where(pd.notnull(df), None)
    df2 = df2.applymap(_safe_value_for_excel)
    return df2

def clean_column_names(columns):
    """
    Replace obvious 'Unnamed' or empty column names with Kolom_i placeholders.
    We'll later drop placeholders that are truly empty.
    """
    new_cols = []
    for i, c in enumerate(columns):
        cstr = "" if c is None else str(c)
        if cstr.strip().lower().startswith("unnamed") or cstr.strip() == "":
            new_cols.append(f"Kolom_{i+1}")
        else:
            new_cols.append(cstr.strip())
    return new_cols

def drop_truly_empty_or_unnamed_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Drop columns that are:
    - Named Kolom_x (generated placeholder) AND
      their entire column values are empty / unwanted / NaN.
    This makes dashboard and export clean while preserving any Kolom_x that actually contain data.
    """
    df2 = df.copy()
    to_drop = []
    for c in df2.columns:
        c_lower = str(c).strip().lower()
        # consider placeholder names
        is_placeholder = c_lower.startswith("kolom_") or c_lower.startswith("unnamed") or c_lower == ""
        if is_placeholder:
            # check if all values are unwanted / null
            all_unwanted = True
            for val in df2[c]:
                if pd.isna(val):
                    continue
                if not is_unwanted_label(val):
                    all_unwanted = False
                    break
            if all_unwanted:
                to_drop.append(c)
    if to_drop:
        df2 = df2.drop(columns=to_drop)
    return df2

# -----------------------
# EXCEL EXPORT
# -----------------------
def download_excel_bytes(df, stats_df=None):
    """
    Create nicely formatted Excel bytes:
    - header bold + colored
    - drop totally-empty placeholder columns (we assume df already cleaned)
    - auto-width, thin borders, wrap text, freeze header
    - include optional stats sheet
    """
    if not isinstance(df, pd.DataFrame):
        raise ValueError("Data yang akan diunduh harus berupa DataFrame.")
    # ensure index reset
    df_clean = safe_for_excel_df(df.reset_index(drop=True))
    stats_clean = safe_for_excel_df(stats_df.reset_index(drop=True)) if stats_df is not None and not stats_df.empty else None

    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Data_Hasil_Edit"

    # styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0077B6", end_color="0077B6", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(border_style="thin", color="000000")
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # write header (clean names)
    for i, col in enumerate(df_clean.columns, start=1):
        colname = str(col).strip()
        cell = ws1.cell(row=1, column=i, value=colname)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # append rows (no header)
    for r in dataframe_to_rows(df_clean, index=False, header=False):
        ws1.append(r)

    # styling content
    for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, min_col=1, max_col=ws1.max_column):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border

    # auto width (cap at 60)
    for col_cells in ws1.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                continue
        adjusted_width = min((max_length + 2), 60)
        ws1.column_dimensions[col_letter].width = adjusted_width

    # freeze header
    ws1.freeze_panes = "A2"

    # statistics sheet
    if stats_clean is not None:
        ws2 = wb.create_sheet("Ringkasan_Statistik")
        for i, col in enumerate(stats_clean.columns, start=1):
            cell = ws2.cell(row=1, column=i, value=str(col).strip())
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        for r in dataframe_to_rows(stats_clean, index=False, header=False):
            ws2.append(r)
        for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=1, max_col=ws2.max_column):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = thin_border
        for col_cells in ws2.columns:
            max_length = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                try:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    continue
            ws2.column_dimensions[col_letter].width = min((max_length + 2), 60)

    # simple charts sheet (non-intrusive)
    ws3 = wb.create_sheet("Grafik_Statistik")
    ws3.append(["Grafik Otomatis (Dihasilkan dari Data)"])
    ws3["A1"].font = Font(bold=True)
    base_row = 3

    # PIE CHART (gender)
    gender_cols = [c for c in df_clean.columns if "gender" in str(c).lower() or "jenis kelamin" in str(c).lower()]
    if gender_cols:
        gcol = gender_cols[0]
        vc = pd.Series(df_clean[gcol].fillna("Unknown").astype(str)).value_counts().reset_index()
        vc.columns = ["Kategori", "Jumlah"]
        ws3.append([])
        ws3.append(["Kategori", "Jumlah"])
        for r in dataframe_to_rows(vc, index=False, header=False):
            ws3.append(r)
        start_row = ws3.max_row - len(vc) + 1
        pie = PieChart()
        pie.title = f"Proporsi {gcol}"
        data_ref = Reference(ws3, min_col=2, min_row=start_row, max_row=start_row + len(vc) - 1)
        labels_ref = Reference(ws3, min_col=1, min_row=start_row, max_row=start_row + len(vc) - 1)
        pie.add_data(data_ref, titles_from_data=False)
        pie.set_categories(labels_ref)
        pie.height = 8
        pie.width = 8
        ws3.add_chart(pie, f"E{base_row}")
        base_row += 20

    # BAR CHART (first categorical)
    cat_cols = df_clean.select_dtypes(include=['object', 'category']).columns.tolist()
    if cat_cols:
        ccol = cat_cols[0]
        vc2 = pd.Series(df_clean[ccol].fillna("Unknown").astype(str)).value_counts().reset_index().head(10)
        vc2.columns = ["Kategori", "Jumlah"]
        ws3.append([])
        ws3.append(["Kategori", "Jumlah"])
        for r in dataframe_to_rows(vc2, index=False, header=False):
            ws3.append(r)
        start_row = ws3.max_row - len(vc2) + 1
        bar = BarChart()
        bar.title = f"Distribusi {ccol}"
        data_ref = Reference(ws3, min_col=2, min_row=start_row, max_row=start_row + len(vc2) - 1)
        labels_ref = Reference(ws3, min_col=1, min_row=start_row, max_row=start_row + len(vc2) - 1)
        bar.add_data(data_ref, titles_from_data=False)
        bar.set_categories(labels_ref)
        bar.height = 10
        bar.width = 12
        ws3.add_chart(bar, f"E{base_row}")
        base_row += 25

        # save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def short_top3_text_from_counts(value_counts: pd.Series) -> str:
    """
    Mengambil 3 nilai teratas dari hasil value_counts dan ubah menjadi teks singkat
    seperti: "Unit A (40), Unit B (32), Unit C (25)".
    """
    try:
        top3 = value_counts.head(3)
        total = int(value_counts.sum())
        parts = [f"{idx} ({val})" for idx, val in top3.items()]
        return ", ".join(parts) + f" — Total: {total}"
    except Exception as e:
        return f"(gagal membuat ringkasan top3: {e})"

    # ================================
# EXPORT PDF FINAL (HD + COVER + TOC + TABEL RAPI)
# ================================
from reportlab.platypus import Table, TableStyle
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
import tempfile, os
from datetime import datetime
import plotly.io as pio
import streamlit as st


def export_dashboard_pdf(figures_dict=None, tables_dict=None, author_name="PPNS Dashboard System", logo_path=None):
    """
    Export PDF gabungan grafik + tabel, lengkap dengan cover dan daftar isi.
    - figures_dict: dict {judul_grafik: fig_plotly}
    - tables_dict: dict {judul_tabel: pd.DataFrame}
    - author_name: nama pembuat laporan (opsional)
    - logo_path: path logo (opsional)
    """

    if not figures_dict and not tables_dict:
        st.warning("⚠ Tidak ada grafik atau tabel yang bisa diekspor.")
        return

    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmpfile:
            c = canvas.Canvas(tmpfile.name, pagesize=A4)
            width, height = A4
            margin = 50
            line_gap = 16

            # --------------------------
            # Fungsi bantu teks rapi
            # --------------------------
            def draw_wrapped_text(c, text, x, y, max_width, font="Helvetica", size=12, line_height=14):
                c.setFont(font, size)
                words = text.split()
                line = ""
                for word in words:
                    test_line = f"{line} {word}".strip()
                    if c.stringWidth(test_line, font, size) > max_width:
                        c.drawString(x, y, line)
                        y -= line_height
                        line = word
                    else:
                        line = test_line
                if line:
                    c.drawString(x, y, line)
                    y -= line_height
                return y

            # ===================================================
            # 📘 1. COVER PAGE
            # ===================================================
            logo_used = None

            # 🔍 Otomatis cari logo.png di folder kerja jika tidak diberikan
            if not logo_path:
                possible_logo = os.path.join(os.getcwd(), "logo.png")
                if os.path.exists(possible_logo):
                    logo_used = possible_logo
            else:
                # Gunakan logo dari argumen (misal hasil upload)
                if os.path.exists(logo_path):
                    logo_used = logo_path
                else:
                    # Coba ubah jadi path absolut
                    abs_path = os.path.join(os.getcwd(), logo_path)
                    if os.path.exists(abs_path):
                        logo_used = abs_path

            # Debug info (bisa dihapus kalau tidak mau tampil)
            st.write("🖼️ Path logo terdeteksi:", logo_used if logo_used else "❌ Tidak ditemukan")

            # Gambar logo di cover jika ditemukan
            if logo_used:
                c.drawImage(logo_used, width/2 - 2*cm, height - 7*cm, width=4*cm, preserveAspectRatio=True)

            c.setFont("Helvetica-Bold", 22)
            c.drawCentredString(width/2, height - 10*cm, "📊 LAPORAN GABUNGAN DASHBOARD PPNS")

            c.setFont("Helvetica", 14)
            c.drawCentredString(width/2, height - 11.5*cm, "Sistem Analisis Data dan Statistik")

            c.setFont("Helvetica", 12)
            c.drawCentredString(width/2, height - 13.5*cm, f"Disusun oleh: {author_name}")
            c.drawCentredString(width/2, height - 14.5*cm,
                                f"Tanggal: {datetime.now().strftime('%d %B %Y, %H:%M:%S')}")
            c.line(margin, height - 15*cm, width - margin, height - 15*cm)
            c.showPage()

            # ===================================================
            # 📄 2. DAFTAR ISI OTOMATIS
            # ===================================================
            c.setFont("Helvetica-Bold", 18)
            c.drawString(margin, height - margin - 30, "📑 Daftar Isi")
            y = height - margin - 60
            c.setFont("Helvetica", 12)

            section_index = 1
            if tables_dict:
                c.drawString(margin, y, "Tabel Data:")
                y -= line_gap
                for title in tables_dict.keys():
                    c.drawString(margin + 20, y, f"{section_index}. {title}")
                    section_index += 1
                    y -= line_gap

            if figures_dict:
                y -= 10
                c.drawString(margin, y, "Grafik Statistik:")
                y -= line_gap
                for title in figures_dict.keys():
                    c.drawString(margin + 20, y, f"{section_index}. {title}")
                    section_index += 1
                    y -= line_gap

            c.showPage()

            # ===================================================
            # 📊 3. ISI LAPORAN (TABEL & GRAFIK)
            # ===================================================
            y = height - margin

            # ---------- TABEL ----------
            if tables_dict:
                for title, df in tables_dict.items():
                    if y < 200:
                        c.showPage()
                        y = height - margin

                    c.setFont("Helvetica-Bold", 14)
                    y = draw_wrapped_text(c, f"📋 {title}", margin, y, width - 2*margin, "Helvetica-Bold", 13)
                    y -= 5

                    data = [df.columns.tolist()] + df.astype(str).values.tolist()

                    table = Table(data, colWidths=[(width - 2*margin) / len(df.columns)] * len(df.columns))
                    style = TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.1, 0.2, 0.5)),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, -1), 9),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                        ('GRID', (0, 0), (-1, -1), 0.3, colors.grey),
                    ])
                    table.setStyle(style)

                    table_height = 15 * len(data)
                    if y - table_height < margin:
                        c.showPage()
                        y = height - margin

                    table.wrapOn(c, width, height)
                    table.drawOn(c, margin, y - table_height)
                    y -= table_height + 25

            # ---------- GRAFIK ----------
            if figures_dict:
                for title, fig in figures_dict.items():
                    if y < 280:
                        c.showPage()
                        y = height - margin
                        c.setFont("Helvetica-Bold", 14)
                        c.drawString(margin, y, "Laporan Dashboard PPNS (Lanjutan)")
                        y -= 40

                    tmp_png = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
                    try:
                        pio.write_image(fig, tmp_png.name, format="png", width=1400, height=900, scale=5)
                        y = draw_wrapped_text(c, f"📈 {title}", margin, y, width - 2*margin, "Helvetica-Bold", 13)
                        y -= 5
                        img_height = 240
                        c.drawImage(tmp_png.name, margin, y - img_height, width=width - 2*margin,
                                    height=img_height, preserveAspectRatio=True)
                        y -= img_height + 25
                    finally:
                        tmp_png.close()
                        os.remove(tmp_png.name)

            # Selesai
            c.save()

            # Tombol download
            with open(tmpfile.name, "rb") as f:
                st.download_button(
                    label="📥 Unduh PDF Laporan Lengkap (HD + Cover + TOC)",
                    data=f.read(),
                    file_name=f"Laporan_PPNS_Final_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
                    mime="application/pdf",
                )

    except Exception as e:
        st.error(f"❌ Gagal mengekspor PDF: {e}")


# ================================
# BACA FILE EXCEL
# ================================
uploaded_file = st.file_uploader("📤 Unggah file Excel (.xlsx)", type=["xlsx"])

if uploaded_file:
    try:
        with st.spinner("⏳ Membaca file & menyiapkan dashboard..."):
            xls = pd.ExcelFile(uploaded_file)
            sheet_names = xls.sheet_names

        # Pilih sheet yang akan dibaca
        selected_sheet = st.selectbox("📄 Pilih Sheet:", sheet_names)

        # Baca file tanpa header untuk deteksi header asli
        raw_df = pd.read_excel(uploaded_file, sheet_name=selected_sheet, header=None, dtype=object)

        if raw_df.empty:
            st.warning(f"⚠ Sheet '{selected_sheet}' kosong.")
            st.stop()

        # ======================================
        # DETEKSI OTOMATIS BARIS HEADER ASLI
        # ======================================
        header_row_idx = None
        for i, row in raw_df.iterrows():
            # cari baris yang mengandung "UNIT KERJA" atau "NO" (biasanya header)
            if row.astype(str).str.contains("UNIT KERJA|NO", case=False, na=False).any():
                header_row_idx = i
                break

        if header_row_idx is not None:
            # gunakan baris yang ditemukan sebagai header
            df = pd.read_excel(uploaded_file, sheet_name=selected_sheet, header=header_row_idx, dtype=object)
        else:
            # fallback jika tidak ditemukan
            df = pd.read_excel(uploaded_file, sheet_name=selected_sheet, header=0, dtype=object)

        # ======================================
        # PEMBERSIHAN DATAFRAME
        # ======================================

        # Hilangkan kolom dengan nama 'Unnamed'
        df.columns = [
            str(c).replace("Unnamed: ", "").strip() if "Unnamed" in str(c) else str(c).strip()
            for c in df.columns
        ]

        # Hapus baris kosong di atas header jika masih ada
        while df.iloc[0].isna().all():
            df = df.iloc[1:].reset_index(drop=True)

        # Hapus baris yang benar-benar kosong
        df.dropna(how="all", inplace=True)
        df.reset_index(drop=True, inplace=True)

        if df.empty:
            st.warning(f"⚠ Sheet '{selected_sheet}' tidak berisi data setelah dibersihkan.")
            st.stop()

        # ======================================
        # SIAPKAN UNTUK UI
        # ======================================
        def drop_truly_empty_or_unnamed_columns(dataframe):
            """Hapus kolom kosong atau unnamed sepenuhnya"""
            return dataframe.loc[:, ~(dataframe.columns.astype(str).str.startswith("Unnamed"))]

        df_for_ui = drop_truly_empty_or_unnamed_columns(df.copy())
        df_for_ui.columns = [str(c).strip() for c in df_for_ui.columns]

        # Ganti NaN jadi tanda '-'
        df_for_ui = df_for_ui.fillna("-")

        # -------------------------------
        # Simpan ke session state
        # -------------------------------
        if "df_original" not in st.session_state:
            st.session_state["df_original"] = df.copy()

        st.session_state["df_current"] = df_for_ui.copy()

        # -------------------------------
        # TAMPILKAN DI UI STREAMLIT
        # -------------------------------
        st.success("✅ File berhasil dibaca dan header diperbaiki otomatis!")
        st.dataframe(df_for_ui, use_container_width=True)

    except Exception as e:
        st.error(f"❌ Terjadi kesalahan saat membaca file: {e}")

    except Exception as e:
        st.error(f"Terjadi kesalahan saat membaca atau memproses file Excel: {e}")
        st.stop()

    # -------------------------------
    # Store original and current in session state
    # -------------------------------
    if "df_original" not in st.session_state:
        st.session_state["df_original"] = df.copy()
    st.session_state["df_current"] = df_for_ui.fillna("-").copy()  # show '-' for missing values in editor

# ================================
# Sidebar style (gradasi biru)
# ================================
st.markdown("""
    <style>
    [data-testid="stSidebar"] {
        background: linear-gradient(to bottom, #003366, #0099ff);
        color: white !important;
    }
    [data-testid="stSidebar"] * {
        color: white !important;
    }
    .sidebar-title {
        font-size: 22px;
        font-weight: bold;
        margin-bottom: 20px;
        color: white;
        text-align: center;
    }
    .sidebar-btn {
        display: block;
        padding: 10px 15px;
        margin: 8px 0;
        border-radius: 8px;
        background-color: rgba(255,255,255,0.1);
        text-decoration: none;
        color: white;
        font-weight: 500;
    }
    .sidebar-btn:hover {
        background-color: rgba(255,255,255,0.3);
        transition: 0.3s;
    }
    </style>
""", unsafe_allow_html=True)

# ================================
# Sidebar menu
# ================================
st.sidebar.markdown('<div class="sidebar-title">📊 Dashboard PPNS</div>', unsafe_allow_html=True)

# Pastikan ada key unik supaya tidak duplicate ID
if "active_tab" not in st.session_state:
    st.session_state["active_tab"] = "📋 Data"

menu = st.sidebar.radio(
    "Navigasi",
    ["📋 Data", "📊 Grafik", "📈 Statistik"],
    index=["📋 Data", "📊 Grafik", "📈 Statistik"].index(st.session_state["active_tab"]),
    label_visibility="collapsed",
    key="sidebar_radio_menu"  # <-- kunci unik untuk menghindari error duplicate ID
)

# Update session_state ketika sidebar diubah
if st.session_state.sidebar_radio_menu != st.session_state.active_tab:
    st.session_state.active_tab = st.session_state.sidebar_radio_menu

# ================================
# Cek apakah file diupload
# ================================
if uploaded_file:
    try:
        # Ambil df_current dari session
        df_current = st.session_state.get("df_current", pd.DataFrame())

        if menu == "📋 Data":
            # === TAB 1 ===
            st.subheader("📋 Tabel Data")
            st.write("🔎 Pencarian cepat — ketik Nama atau NIP, lalu pilih")

            # heuristik kolom nama/nip
            name_cols = [c for c in df_current.columns if "nama" in str(c).lower()]
            nip_cols = [c for c in df_current.columns if "nip" in str(c).lower()]
            primary_name_col = name_cols[0] if name_cols else (df_current.columns[0] if len(df_current.columns) > 0 else None)
            primary_nip_col = nip_cols[0] if nip_cols else None

            options = []
            if primary_name_col and primary_nip_col:
                for _, row in df_current[[primary_name_col, primary_nip_col]].fillna("-").iterrows():
                    nm, npv = str(row[primary_name_col]), str(row[primary_nip_col])
                    display_opt = f"{nm} — {npv}"
                    options.append(display_opt)
            else:
                for val in df_current.iloc[:, 0].fillna("-").astype(str).tolist():
                    options.append(val)

            selected_search = st.selectbox("Pilih dari daftar:", options) if options else st.text_input("Cari Nama / NIP:")
            if selected_search:
                matched_rows = pd.DataFrame()
                if "—" in str(selected_search) and primary_nip_col:
                    parts = [p.strip() for p in selected_search.split("—")]
                    if len(parts) >= 2:
                        left, right = parts[0], parts[-1]
                        matched_rows = df_current[df_current[primary_nip_col].astype(str).str.strip() == right]
                        if matched_rows.empty and primary_name_col:
                            matched_rows = df_current[df_current[primary_name_col].astype(str).str.contains(left, case=False, na=False)]
                else:
                    q = str(selected_search).strip()
                    mask = pd.Series(False, index=df_current.index)
                    for c in df_current.columns:
                        try:
                            mask |= df_current[c].astype(str).str.contains(q, case=False, na=False)
                        except:
                            continue
                    matched_rows = df_current[mask]

                if not matched_rows.empty:
                    first = matched_rows.iloc[0]
                    cols_left, cols_right = st.columns([2, 3])
                    with cols_left:
                        st.markdown("### 🔖 Detail Terpilih")
                        for k in df_current.columns:
                            val = first.get(k, "")
                            if val is None or is_unwanted_label(val):
                                continue
                            st.markdown(f"<div><b>{k}:</b> {val}</div>", unsafe_allow_html=True)
                    with cols_right:
                        st.write("Ringkasan Data")
                        display_summary = matched_rows.copy()
                        for c in display_summary.columns:
                            display_summary[c] = display_summary[c].where(~display_summary[c].astype(str).str.strip().str.lower().isin(UNWANTED_LABELS), pd.NA)
                        st.write(display_summary.head(10).fillna("-"))
                        if len(matched_rows) > 1:
                            st.info(f"{len(matched_rows)} hasil cocok. Menampilkan 10 teratas.")
                else:
                    st.warning("Tidak ditemukan data.")

            # show editable table (cleaned)
            st.data_editor(df_current, num_rows="dynamic", use_container_width=True, height=520)

            c1, c2, c3, c4 = st.columns(4)
            with c1:
                if st.button("🔁 Reset ke Asli"):
                    df_orig = st.session_state.get("df_original", pd.DataFrame()).copy()
                    df_ui = drop_truly_empty_or_unnamed_columns(df_orig.copy())
                    st.session_state["df_current"] = df_ui.fillna("-").copy()
                    st.experimental_rerun()
            with c2:
                if st.button("🔄 Sinkron ke Original"):
                    edited = df_current.replace("-", pd.NA).copy()
                    df_orig = st.session_state.get("df_original", pd.DataFrame()).copy()
                    for c in edited.columns:
                        if c in df_orig.columns:
                            df_orig.loc[edited.index, c] = edited[c].values
                    for c in edited.columns:
                        if c not in df_orig.columns:
                            df_orig[c] = pd.NA
                            df_orig.loc[edited.index, c] = edited[c].values
                    st.session_state["df_original"] = df_orig
                    st.success("Sukses menyinkronkan.")
            with c3:
                stats_summary = pd.DataFrame()
                num_cols = df_current.select_dtypes(include=["number"]).columns.tolist()
                if num_cols:
                    stats_summary = df_current[num_cols].describe().transpose().reset_index().rename(columns={"index": "Kolom"})
                df_for_export = df_current.replace("-", pd.NA).copy()
                df_for_export.columns = [str(c).strip() for c in df_for_export.columns]
                excel_bytes = download_excel_bytes(df_for_export, stats_summary)
                st.download_button(
                    label="💾 Unduh Hasil Edit + Statistik",
                    data=excel_bytes,
                    file_name=f"hasil_edit_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            with c4:
                if st.button("📌 Simpan ke Session"):
                    st.session_state["snapshot"] = df_current.replace("-", pd.NA).copy()
                    st.success("Snapshot disimpan.")

            # === GRAFIK GABUNGAN ===
            st.markdown("---")
            st.subheader("📊 Grafik Gabungan (Semua Sheet)")
            export_figs = {}

            all_dfs = []
            for sheet in sheet_names:
                try:
                    tmp_raw = pd.read_excel(uploaded_file, sheet_name=sheet, header=None, dtype=object)
                    header_row_tmp = detect_header_row(tmp_raw)
                    df_tmp = pd.read_excel(uploaded_file, sheet_name=sheet, header=header_row_tmp, dtype=object)
                    df_tmp.columns = clean_column_names(df_tmp.columns)
                    df_tmp.dropna(how="all", inplace=True)
                    if not df_tmp.empty:
                        df_tmp = df_tmp.copy()
                        df_tmp["sheetname"] = sheet
                        all_dfs.append(df_tmp)
                except Exception:
                    continue

            if all_dfs:
                df_all = pd.concat(all_dfs, ignore_index=True)
                df_all.columns = clean_column_names(df_all.columns)

                # Bar chart per sheet
                count_per_sheet = df_all["sheetname"].value_counts().reset_index()
                count_per_sheet.columns = ["Sheet", "Jumlah Data"]
                fig_sheet = px.bar(
                    count_per_sheet,
                    x="Sheet",
                    y="Jumlah Data",
                    text="Jumlah Data",
                    title="Jumlah Baris Data per Sheet (Gabungan Semua Sheet)"
                )
                fig_sheet.update_layout(xaxis_tickangle=-45)
                st.plotly_chart(fig_sheet, use_container_width=True)
                sheet_text = ", ".join([f"{r['Sheet']}: {r['Jumlah Data']}" for _, r in count_per_sheet.iterrows()])
                st.markdown(f"Per-sheet: {sheet_text}.")
                export_figs["Jumlah Baris per Sheet"] = fig_sheet

                # Unit kerja top
                unit_cols = [c for c in df_all.columns if "unit" in str(c).lower() or "kerja" in str(c).lower() or "unitkerja" in str(c).lower()]
                if unit_cols:
                    uc = unit_cols[0]
                    cleaned_uc = clean_series_for_vc(df_all[uc])
                    if not cleaned_uc.empty:
                        top_unit = cleaned_uc.value_counts().reset_index().head(10)
                        top_unit.columns = ["Unit Kerja", "Jumlah"]
                        fig_unit = px.bar(
                            top_unit,
                            x="Unit Kerja",
                            y="Jumlah",
                            text="Jumlah",
                            title="Top 10 Unit Kerja (Gabungan Semua Sheet)"
                        )
                        fig_unit.update_layout(xaxis_tickangle=-45)
                        st.plotly_chart(fig_unit, use_container_width=True)
                        top3_text = short_top3_text_from_counts(cleaned_uc.value_counts())
                        st.markdown(f"Top 3 Unit Kerja: {top3_text}.")
                        export_figs["Top Unit Kerja"] = fig_unit

                # Gender
                gender_keys_all = [c for c in df_all.columns if "gender" in str(c).lower() or "jenis kelamin" in str(c).lower()]
                if gender_keys_all:
                    gcol_all = gender_keys_all[0]
                    cleaned_g = clean_series_for_vc(df_all[gcol_all])
                    if not cleaned_g.empty:
                        tmpg = cleaned_g.value_counts().reset_index()
                        tmpg.columns = ["Kategori", "Jumlah"]
                        figg = px.pie(tmpg, names="Kategori", values="Jumlah", title=f"Proporsi {gcol_all} (Gabungan Semua Sheet)", hole=0.35)
                        figg.update_traces(textinfo="label+percent", hovertemplate="%{label}: %{value} orang")
                        st.plotly_chart(figg, use_container_width=True)
                        vc = cleaned_g.value_counts()
                        total = int(vc.sum())
                        parts = [f"{label}: {count} ({count/total*100:.1f}%)" for label, count in vc.items()]
                        st.markdown("" + ", ".join(parts) + f", Total: {total}.")
                        export_figs[f"Proporsi {gcol_all}"] = figg

                # Numeric histogram
                num_cols_all = df_all.select_dtypes(include='number').columns.tolist()
                if not num_cols_all:
                    for c in df_all.columns:
                        try:
                            conv = pd.to_numeric(df_all[c], errors="coerce")
                            if conv.dropna().shape[0] > 0:
                                num_cols_all.append(c)
                                break
                        except Exception:
                            continue

                if num_cols_all:
                    ncol = num_cols_all[0]
                    nums = pd.to_numeric(df_all[ncol], errors="coerce").dropna()
                    if not nums.empty:
                        fig_hist = px.histogram(nums, x=nums, nbins=15, title=f"Distribusi Numerik: {ncol} (Gabungan Semua Sheet)")
                        st.plotly_chart(fig_hist, use_container_width=True)
                        st.markdown(f"{ncol} — count: {len(nums)}, mean: {nums.mean():.2f}, median: {nums.median():.2f}.")
                        export_figs[f"Distribusi {ncol}"] = fig_hist
            else:
                st.info("Tidak ditemukan sheet berisi data valid untuk dibuat grafik gabungan.")

            # === PDF EXPORT BUTTON ===
            st.markdown("### 📄 Ekspor Laporan PDF (Grafik Gabungan)")
            if st.button("⬇ Download Laporan Gabungan (PDF)"):
                export_dashboard_pdf(export_figs)

        elif menu == "📊 Grafik":
            # === TAB 2 ===
            st.subheader("📊 Visualisasi Otomatis")
            for_display_df = df_current.replace("-", pd.NA).copy()

            gender_keys = [c for c in for_display_df.columns if "gender" in str(c).lower() or "jenis kelamin" in str(c).lower()]
            if gender_keys:
                gcol = gender_keys[0]
                cleaned = clean_series_for_vc(for_display_df[gcol])
                if not cleaned.empty:
                    tmp = cleaned.value_counts().reset_index()
                    tmp.columns = ["Kategori", "Jumlah"]
                    fig = px.pie(tmp, names="Kategori", values="Jumlah", title=f"Proporsi {gcol}", hole=0.35)
                    fig.update_traces(textinfo="label+percent", hovertemplate="%{label}: %{value} orang")
                    st.plotly_chart(fig, use_container_width=True)
                    vc = cleaned.value_counts()
                    total = int(vc.sum())
                    parts = [f"{label}: {count} ({count/total*100:.1f}%)" for label, count in vc.items()]
                    st.markdown("" + ", ".join(parts) + f", Total: {total}.")

            for col in for_display_df.columns:
                try:
                    if for_display_df[col].dtype == object or for_display_df[col].dtype.name == "category":
                        vc_series = clean_series_for_vc(for_display_df[col])
                        if not vc_series.empty:
                            vc = vc_series.value_counts().reset_index().head(20)
                            vc.columns = ["Kategori", "Jumlah"]
                            bar = px.bar(vc, x="Kategori", y="Jumlah", text="Jumlah", title=f"Distribusi: {col}")
                            bar.update_layout(xaxis_tickangle=-45)
                            st.plotly_chart(bar, use_container_width=True)
                            counts = vc_series.value_counts()
                            if not counts.empty:
                                top3_text = short_top3_text_from_counts(counts)
                                st.markdown(f"Top 3 {col}: {top3_text}.")
                    elif pd.api.types.is_numeric_dtype(for_display_df[col]):
                        nums = pd.to_numeric(for_display_df[col], errors="coerce").dropna()
                        if not nums.empty:
                            hist = px.histogram(nums, x=nums, nbins=12, title=f"Distribusi Numerik: {col}")
                            st.plotly_chart(hist, use_container_width=True)
                            st.markdown(f"{col} — count: {len(nums)}, mean: {nums.mean():.2f}, median: {nums.median():.2f}.")
                except Exception as e:
                    st.warning(f"Gagal membuat grafik {col}: {e}")

        elif menu == "📈 Statistik":
            # === TAB 3 ===
            st.subheader("📈 Ringkasan Statistik")
            current_df = df_current.replace("-", pd.NA)
            num_cols = current_df.select_dtypes(include=["number"]).columns.tolist()
            if num_cols:
                st.markdown("Statistik Numerik")
                st.dataframe(current_df[num_cols].describe().transpose())

            cat_cols = current_df.select_dtypes(include=["object", "category"]).columns.tolist()
            if cat_cols:
                st.markdown("Top Kategori")
                for c in cat_cols:
                    cleaned_ser = clean_series_for_vc(current_df[c])
                    unique_count = cleaned_ser.nunique()
                    if unique_count == 0:
                        continue
                    st.write(f"{c}** — unik (bersih): {unique_count} item")
                    vc = cleaned_ser.value_counts().head(10).rename_axis(c).reset_index().rename(columns={"index": c, 0: "Jumlah"})
                    vc.columns = [c, "Jumlah"]
                    st.dataframe(vc)
                    counts = cleaned_ser.value_counts()
                    if not counts.empty:
                        top3_text = short_top3_text_from_counts(counts)
                        st.markdown(f"Top 3 {c}: {top3_text}.")
            st.markdown("---")
            st.markdown("<footer>© 2025 Kementerian Kelautan dan Perikanan — Dashboard PPNS</footer>", unsafe_allow_html=True)

    except Exception as e:
        st.error(f"Terjadi kesalahan saat membaca file: {e}")
else:
    st.info("📥 Silakan upload file Excel terlebih dahulu.")


# -----------------------
# UTIL FUNCTIONS USED ABOVE (kehilangan fungsi helper kecil ada di bawah)
# -----------------------
def short_top3_text_from_counts(counts: pd.Series) -> str:
    top = counts.head(3)
    items = [f"{idx} ({val})" for idx, val in top.items()]
    return ", ".join(items)
